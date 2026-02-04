"""Helpers para geração de relatórios em XLSX usando openpyxl."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def gerar_xlsx_tabela(titulo, colunas, linhas):
    """
    Gera um arquivo XLSX com uma tabela.

    Args:
        titulo: Título do relatório (usado como nome da planilha)
        colunas: Lista de nomes das colunas
        linhas: Lista de listas (cada lista é uma linha)

    Returns:
        bytes: Conteúdo do XLSX
    """
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31] if len(titulo) > 31 else titulo  # Excel limita a 31 chars

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Cabeçalho
    for col_idx, coluna in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=coluna)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Dados
    for row_idx, linha in enumerate(linhas, start=2):
        for col_idx, valor in enumerate(linha, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Ajustar largura das colunas
    for col_idx in range(1, len(colunas) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
